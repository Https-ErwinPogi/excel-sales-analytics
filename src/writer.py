from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart


def create_report(
        summary,
        region_sales,
        product_sales,
        customer_sales,
        top_product,
        top_region,
        avg_delivery_time,
        top_products):

    # Create Workbook
    wb = Workbook()

    # =========================
    # SUMMARY SHEET
    # =========================
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws['A1'] = "SALES SUMMARY REPORT"

    # Sales Overview
    ws['A3'] = "Total Sales"
    ws['B3'] = summary

    # Operations
    ws['A4'] = "Average Delivery Time"
    ws['B4'] = avg_delivery_time

    # Performances
    top_product_name, top_product_value = top_product
    top_region_name, top_region_value = top_region

    ws['A6'] = "Top Product"
    ws['B6'] = top_product_name
    ws['C6'] = top_product_value

    ws['A7'] = "Top Region"
    ws['B7'] = top_region_name
    ws['C7'] = top_region_value

    # =========================
    # SALES BY REGION
    # =========================
    ws_region = wb.create_sheet(title="Sales by Region")

    ws_region.append(["Region", "Total Sales"])

    for region, sales in region_sales.items():
        ws_region.append([region, sales])

    # =========================
    # PRODUCT ANALYSIS
    # =========================
    ws_product = wb.create_sheet(title="Product Analysis")

    ws_product.append(["Product", "Total Sales"])

    for product, sales in product_sales.items():
        ws_product.append([product, sales])

    # =========================
    # CUSTOMER ANALYSIS
    # =========================
    ws_customer = wb.create_sheet(title="Customer Analysis")

    ws_customer.append(["Customer", "Total Sales"])

    for customer, sales in customer_sales.items():
        ws_customer.append([customer, sales])

    # =========================
    # TOP PRODUCTS PER REGION
    # =========================

    ws_top = wb.create_sheet(title="Top Products")

    ws_top.append(["Region", "Product", "Sales"])

    # expects: data["top_products"] = {region: [(product, sales), ...]}
    for region, products in top_products.items():
        for product, sales in products:
            ws_top.append([region, product, sales])

    # =========================
    # SAVE FILE
    # =========================
    wb.save("output/report.xlsx")
