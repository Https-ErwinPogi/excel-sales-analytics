from openpyxl import load_workbook


def clean_value(value):
    if value is None:
        return "Unknown"
    return value


def load_data(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Step 1: headers
    headers = [cell.value for cell in ws[1]]

    data = []

    # Step 2: loop rows
    for row in ws.iter_rows(min_row=2, values_only=True):

        # Step 3: clean row values
        cleaned_row = [clean_value(value) for value in row]

        # Step 4: convert to dictionary
        row_dict = dict(zip(headers, cleaned_row))

        # Step 5: store result
        data.append(row_dict)

    return data
